import re
import sys
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QFileDialog, QComboBox, QSpinBox, QMessageBox, QLineEdit, QListWidget,
    QListWidgetItem, QGroupBox
)

import os
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase import pdfmetrics

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def resource_path(relative_path: str) -> str:
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative_path)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, relative_path)

# ----------------------------
# 1) 列名自动识别（中/英版）
# ----------------------------
COLMAPS = [
    {"sku": "Sku", "cn": "产品名称(中)", "en": "产品名称(英)", "bin": "库位", "qty": "数量"},  # 中文版
    {"sku": "SKU", "cn": "Product Name (Cn)", "en": "Product Name (En)", "bin": "Bin", "qty": "Qty."},  # 英文版
]

def detect_columns(df: pd.DataFrame) -> Dict[str, str]:
    cols = set(df.columns.astype(str))
    for m in COLMAPS:
        if all(m[k] in cols for k in ("sku", "cn", "en", "bin", "qty")):
            return m

    # 兼容大小写/空格差异
    norm = {c: re.sub(r"\s+", " ", str(c)).strip().lower() for c in df.columns}

    def find(name: str) -> Optional[str]:
        target = name.strip().lower()
        for orig, n in norm.items():
            if n == target:
                return orig
        return None

    for m in COLMAPS:
        cand = {}
        ok = True
        for k, v in m.items():
            f = find(v)
            if not f:
                ok = False
                break
            cand[k] = f
        if ok:
            return cand

    raise ValueError("无法自动识别列名：请确认是WMS导出的中/英文库存明细表。")


# ----------------------------
# 2) 库位清洗：只支持 Freezer(C,D)
# ----------------------------
FREEZER_PREFIX = "Freezer(C,D)"

def clean_bin(raw: str) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw)
    if not s.startswith(FREEZER_PREFIX):
        return None
    s = s[len(FREEZER_PREFIX):]
    s = re.sub(r"\(.*?\)", "", s)
    s = re.sub(r"（.*?）", "", s)
    return s.strip() or None

# ----------------------------
# 3) C区范围生成 + 排序（锁定规则）
# C{排}-{行}-{层}
# 配对：(1,2)(3,4)...；靠过道排：k%4==1 -> A=k+1；k%4==3 -> A=k
# 输出：排对(小->大) -> 层(小->大) -> 行两两 -> A组 -> B组（若B不在用户范围则跳过）
# ----------------------------
def c_aisle_order(p_min: int, p_max: int, r_min: int, r_max: int, l_min: int, l_max: int) -> List[str]:
    def max_row_for_c(p: int) -> int:
        return min(r_max, 9) if 3 <= p <= 14 else r_max

    bins: List[str] = []
    for k in range(1, 24, 2):
        p1, p2 = k, k + 1
        A, B = (p2, p1) if (k % 4 == 1) else (p1, p2)

        r_max_A = max_row_for_c(A)
        r_max_B = max_row_for_c(B)
        r_max_pair = max(r_max_A, r_max_B)

        r = r_min
        while r <= r_max_pair:
            r1, r2 = r, r + 1

            for l in range(l_min, l_max + 1):
                if p_min <= A <= p_max and r1 <= r_max_A:
                    bins.append(f"C{A}-{r1}-{l}")
                    if r2 <= r_max_A:
                        bins.append(f"C{A}-{r2}-{l}")

                if p_min <= B <= p_max and r1 <= r_max_B:
                    bins.append(f"C{B}-{r1}-{l}")
                    if r2 <= r_max_B:
                        bins.append(f"C{B}-{r2}-{l}")

            r += 2

    return bins

# ----------------------------
# 4) D区范围生成 + 排序
# D-{排}B-{层}；排序：层升序 -> 排升序
# D排：1-35；层：1-6
# ----------------------------
def d_order(p_min: int, p_max: int, l_min: int, l_max: int) -> List[str]:
    bins = []
    for l in range(l_min, l_max + 1):
        for p in range(p_min, p_max + 1):
            bins.append(f"D-{p}B-{l}")
    return bins

# ----------------------------
# 4.5) 全仓库位顺序索引缓存（避免每次重建）
# ----------------------------
_C_INDEX: Optional[Dict[str, int]] = None
_D_INDEX: Optional[Dict[str, int]] = None

def get_bin_index() -> Tuple[Dict[str, int], Dict[str, int]]:
    """
    返回 (C区索引, D区索引)，只在第一次调用时生成一次。
    注意：这里用的是“全仓范围”，用于排序，而不是用户选择范围。
    """
    global _C_INDEX, _D_INDEX
    if _C_INDEX is None:
        _C_INDEX = {b: i for i, b in enumerate(c_aisle_order(1, 24, 1, 48, 1, 6))}
    if _D_INDEX is None:
        _D_INDEX = {b: i for i, b in enumerate(d_order(1, 35, 1, 6))}
    return _C_INDEX, _D_INDEX

# ----------------------------
# 5) 数据结构
# ----------------------------
@dataclass
class Row:
    cn: str
    en: str
    bin: str
    qty: str


# ----------------------------
# 6) 构建范围盘点行：补空 + 多行
# - 多行：同库位多条记录按多行输出
# - 空库位：数量=0cs，品名留空
# ----------------------------
def build_rows_for_range(df: pd.DataFrame, col: Dict[str, str], target_bins: List[str], zone: str) -> List[Row]:
    tmp = df[[col["sku"], col["cn"], col["en"], col["bin"], col["qty"]]].copy()
    tmp["_bin_clean"] = tmp[col["bin"]].apply(clean_bin)
    tmp = tmp[tmp["_bin_clean"].notna()]

    # 区域过滤
    if zone == "C":
        tmp = tmp[tmp["_bin_clean"].str.match(r"^C\d+-\d+-\d+$", na=False)]
    elif zone == "D":
        tmp = tmp[tmp["_bin_clean"].str.match(r"^D-\d+B-\d+$", na=False)]

    groups: Dict[str, List[Tuple[str, str, str]]] = {}
    for _, r in tmp.iterrows():
        b = str(r["_bin_clean"])
        cn = "" if pd.isna(r[col["cn"]]) else str(r[col["cn"]])
        en = "" if pd.isna(r[col["en"]]) else str(r[col["en"]])
        qty_raw = "" if pd.isna(r[col["qty"]]) else str(r[col["qty"]])
        groups.setdefault(b, []).append((cn, en, qty_raw))

    out: List[Row] = []
    for b in target_bins:
        if b in groups:
            for cn, en, qty_raw in groups[b]:
                out.append(Row(cn=cn, en=en, bin=b, qty=qty_raw))
        else:
            out.append(Row(cn="", en="", bin=b, qty="0cs"))
    return out


# ----------------------------
# 7) 构建按产品盘点行：不补空
# - 只输出选中SKU实际出现的库位
# - 同库位多条记录按多行输出
# - 输出按仓库顺序排序（C按c_aisle_order全仓索引，D按d_order全仓索引）
# ----------------------------
def build_rows_for_products(df: pd.DataFrame, col: Dict[str, str], selected_skus: List[str], zone: str) -> List[Row]:
    if not selected_skus:
        return []

    tmp = df[[col["sku"], col["cn"], col["en"], col["bin"], col["qty"]]].copy()
    tmp["_bin_clean"] = tmp[col["bin"]].apply(clean_bin)
    tmp = tmp[tmp["_bin_clean"].notna()]

    # SKU过滤
    tmp["_sku"] = tmp[col["sku"]].astype(str)
    selected_skus = [str(s) for s in selected_skus]
    tmp = tmp[tmp["_sku"].isin(selected_skus)].copy()

    # 区域过滤：C / D / ALL
    if zone == "C":
        tmp = tmp[tmp["_bin_clean"].astype(str).str.match(r"^C\d+-\d+-\d+$")]
    elif zone == "D":
        tmp = tmp[tmp["_bin_clean"].astype(str).str.match(r"^D-\d+B-\d+$")]

    c_index, d_index = get_bin_index()

    def bin_rank(bin_code: str) -> int:
        if bin_code.startswith("C"):
            return c_index.get(bin_code, 10**9)
        if bin_code.startswith("D-"):
            return d_index.get(bin_code, 10**9)
        return 10**9

    out: List[Row] = []

    # 关键：按用户选择顺序输出 SKU 分组
    for sku in selected_skus:
        sku_df = tmp[tmp["_sku"] == sku].copy()

        # 每个 SKU 内按库位盘点顺序排序
        sku_df["_rank"] = sku_df["_bin_clean"].astype(str).map(bin_rank)
        sku_df.sort_values(by=["_rank", "_bin_clean"], inplace=True)

        for _, r in sku_df.iterrows():
            b = str(r["_bin_clean"])
            cn = "" if pd.isna(r[col["cn"]]) else str(r[col["cn"]])
            en = "" if pd.isna(r[col["en"]]) else str(r[col["en"]])
            qty_raw = "" if pd.isna(r[col["qty"]]) else str(r[col["qty"]])
            out.append(Row(cn=cn, en=en, bin=b, qty=qty_raw))

    return out

_CJK_FONT_CACHE: Optional[str] = None

def setup_cjk_font() -> str:
    """
    返回可用的中文字体名（已注册到 reportlab）。
    只注册一次：用 module-level cache 避免重复 registerFont。
    优先：微软雅黑 -> 宋体 -> CID(STSong-Light)
    """
    global _CJK_FONT_CACHE
    if _CJK_FONT_CACHE:
        return _CJK_FONT_CACHE

    registered = set(pdfmetrics.getRegisteredFontNames())

    # 1) Windows 常见字体路径
    candidates = [
        (r"C:\Windows\Fonts\msyh.ttf", "MSYH"),          # 优先用 TTF（TTC 兼容性更不稳定）
        (r"C:\Windows\Fonts\msyh.ttc", "MSYH_TTC"),
        (r"C:\Windows\Fonts\msyhbd.ttc", "MSYHBD_TTC"),
        (r"C:\Windows\Fonts\simsun.ttc", "SIMSUN_TTC"),
        (r"C:\Windows\Fonts\simhei.ttf", "SIMHEI"),
        (r"C:\Windows\Fonts\arialuni.ttf", "ARIALUNI"),
    ]

    for path, name in candidates:
        if not os.path.exists(path):
            continue

        # 如果同名已经注册过，直接复用
        if name in registered:
            _CJK_FONT_CACHE = name
            return name

        try:
            pdfmetrics.registerFont(TTFont(name, path))
            _CJK_FONT_CACHE = name
            return name
        except Exception:
            # 尝试下一个候选
            continue

    # 2) CID fallback（同样避免重复注册）
    cid_name = "STSong-Light"
    if cid_name not in registered:
        pdfmetrics.registerFont(UnicodeCIDFont(cid_name))

    _CJK_FONT_CACHE = cid_name
    return cid_name

# ----------------------------
# 8) 导出 PDF（A4、自动分页、列宽自适应、页脚生成时间）
# ----------------------------
def export_pdf(rows: List[Row], filepath: str, title: str):
    CN_FONT = setup_cjk_font()

    styles = getSampleStyleSheet()
    styleN = styles["Normal"]
    styleN.fontName = CN_FONT
    styleN.fontSize = 9

    styleH = styles["Heading2"]
    styleH.fontName = CN_FONT
    styleH.fontSize = 14

    doc = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=36)

    data = [["产品名称中文", "产品名称英文", "库位", "数量"]]
    for r in rows:
        data.append([r.cn, r.en, r.bin, r.qty])

    page_w, _ = A4
    usable_w = page_w - doc.leftMargin - doc.rightMargin

    def max_width(col_idx: int) -> float:
        mx = 0.0
        for row in data:
            s = "" if row[col_idx] is None else str(row[col_idx])
            w = stringWidth(s,  CN_FONT, 9) + 12
            mx = max(mx, w)
        return mx

    widths = [max_width(0), max_width(1), max_width(2), max_width(3)]

    # 给库位/数量最低宽度
    widths[2] = max(widths[2], 120)
    widths[3] = max(widths[3], 60)

    total = sum(widths)
    if total > usable_w:
        fixed = widths[2] + widths[3]
        remain = max(usable_w - fixed, 80)
        w0, w1 = widths[0], widths[1]
        s = max(w0 + w1, 1)
        widths[0] = remain * (w0 / s)
        widths[1] = remain * (w1 / s)

    table_data = []
    for row in data:
        table_data.append([
            Paragraph(str(row[0]), styleN),
            Paragraph(str(row[1]), styleN),
            Paragraph(str(row[2]), styleN),
            Paragraph(str(row[3]), styleN),
        ])

    tbl = Table(table_data, colWidths=widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), CN_FONT),
        ("FONTNAME", (0, 0), (-1, -1), CN_FONT),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))

    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def on_page(canvas, _doc):
        canvas.saveState()
        canvas.setFont(CN_FONT, 8)
        canvas.drawRightString(page_w - doc.rightMargin, 18, f"Generated: {gen_time}")
        canvas.restoreState()

    story = [
        Paragraph(title, styleH),
        Spacer(1, 8),
        tbl,
    ]
    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)

def export_excel(rows: List[Row], filepath: str, title: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # 第一行：标题
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 第二行：表头
    headers = ["产品名称中文", "产品名称英文", "库位", "数量"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据从第 3 行开始
    for i, r in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=r.cn)
        ws.cell(row=i, column=2, value=r.en)
        ws.cell(row=i, column=3, value=r.bin)
        ws.cell(row=i, column=4, value=r.qty)

    # 冻结标题+表头
    ws.freeze_panes = "A3"

    # 自动列宽（简单版）
    for col in range(1, 5):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    wb.save(filepath)

# ----------------------------
# 9) 最小GUI（库位范围盘点 + 按产品盘点）
# ----------------------------
class Main(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("盘库表生成器 (Excel/PDF)")
        self.setWindowIcon(QIcon(resource_path("icon_stock_count.ico")))
        self.excel_path: Optional[str] = None
        self.df: Optional[pd.DataFrame] = None
        self.col: Optional[Dict[str, str]] = None

        font = self.font()
        font.setPointSize(font.pointSize() + 3)
        self.setFont(font)

        layout = QVBoxLayout()

        # 选择文件
        file_row = QHBoxLayout()
        self.file_label = QLabel("未选择Excel")
        btn_pick = QPushButton("选择Excel")
        btn_pick.clicked.connect(self.pick_excel)
        file_row.addWidget(self.file_label, 1)
        file_row.addWidget(btn_pick)
        layout.addLayout(file_row)

        # 模式 + 区域
        top_row = QHBoxLayout()
        self.mode = QComboBox()
        self.mode.addItems(["库位范围盘点", "按产品盘点"])
        self.zone = QComboBox()
        self.zone.addItems(["C区", "D区"])  # 范围模式默认；产品模式会动态加“C区+D区”
        top_row.addWidget(QLabel("模式:"))
        top_row.addWidget(self.mode)
        top_row.addSpacing(12)
        top_row.addWidget(QLabel("区域:"))
        top_row.addWidget(self.zone)
        self.export_type = QComboBox()
        self.export_type.addItems(["PDF", "Excel"])
        top_row.addSpacing(12)
        top_row.addWidget(QLabel("导出:"))
        top_row.addWidget(self.export_type)
        top_row.addStretch(1)
        layout.addLayout(top_row)
        # 让区域下拉框自动扩展宽度，避免显示成 C...区
        self.zone.setSizeAdjustPolicy(QComboBox.AdjustToContents)
        self.zone.setMinimumContentsLength(7)   # 至少能放下 “C区+D区”
        self.zone.setMinimumWidth(130)          # 你可以改成 150 更稳
        self.zone.view().setMinimumWidth(160)   # 下拉列表宽度也加大

        # 说明区（Instructions）
        self.grp_help = QGroupBox("说明")
        help_layout = QVBoxLayout()
        help_layout.setContentsMargins(12, 10, 12, 10)

        help_text = QLabel(
            "使用顺序：\n"
            "1) 在 WMS 的「库存详情」导出 Excel。\n"
            "2) 点击右上角「选择Excel」并选中导出的文件。\n"
            "3) 选择模式（库位范围盘点 / 按产品盘点），区域（C区 / D区 / C区+D区）和输出格式（PDF/Excel）。\n"
            "4) 设置范围或勾选SKU后，点击「生成文件」。\n\n"
            "库位顺序说明（拿到盘库表如何去盘）：\n"
            "• C区：按双深货架盘点逻辑生成顺序——每次以相邻4个库位为一组，先按层从低到高逐层往上，"
            "再水平移动到下一组行位；盘完两排后推进到下面两排，如：(1,2)→(3,4)…\n"
            "• D区：按层从低到高，一层一层盘；每层内按排号从小到大（D-1B-N层 → D-2B-N层 → …）。"
        )
        help_text.setWordWrap(True)
        help_text.setTextInteractionFlags(Qt.TextSelectableByMouse)  # 可选中复制
        help_layout.addWidget(help_text)

        self.grp_help.setLayout(help_layout)
        layout.addWidget(self.grp_help)

        # 库位范围盘点组
        self.grp_range = QGroupBox("库位范围参数")
        rng_layout = QHBoxLayout()
        rng_layout.setContentsMargins(12, 10, 12, 10)
        rng_layout.setSpacing(18)  # 三个分区之间的间距（均匀）

        # SpinBoxes
        self.p_min = QSpinBox(); self.p_min.setRange(1, 35); self.p_min.setValue(1)
        self.p_max = QSpinBox(); self.p_max.setRange(1, 35); self.p_max.setValue(24)

        self.r_min = QSpinBox(); self.r_min.setRange(1, 48); self.r_min.setValue(1)
        self.r_max = QSpinBox(); self.r_max.setRange(1, 48); self.r_max.setValue(48)

        self.l_min = QSpinBox(); self.l_min.setRange(1, 6); self.l_min.setValue(1)
        self.l_max = QSpinBox(); self.l_max.setRange(1, 6); self.l_max.setValue(6)

        def make_range_group(label_text: str, a: QSpinBox, b: QSpinBox) -> QWidget:
            w = QWidget()
            h = QHBoxLayout(w)
            h.setContentsMargins(0, 0, 0, 0)
            h.setSpacing(6)  # 组内间距（小且均匀）

            lbl = QLabel(label_text)
            lbl.setMinimumWidth(60)  # 让“排/行/层”对齐，看起来像三分区
            tilde = QLabel("~")
            tilde.setAlignment(Qt.AlignCenter)

            h.addWidget(lbl)
            h.addWidget(a)
            h.addWidget(tilde)
            h.addWidget(b)
            return w

        grp_p = make_range_group("排：", self.p_min, self.p_max)
        grp_r = make_range_group("行(仅C区)：", self.r_min, self.r_max)
        grp_l = make_range_group("层：", self.l_min, self.l_max)

        rng_layout.addWidget(grp_p)
        rng_layout.addWidget(grp_r)
        rng_layout.addWidget(grp_l)
        rng_layout.addStretch(1)  # 右侧留白

        self.grp_range.setLayout(rng_layout)
        layout.addWidget(self.grp_range)

        # 按产品盘点组
        self.grp_prod = QGroupBox("按产品盘点")
        prod_layout = QVBoxLayout()

        self.search_timer = QTimer(self)
        self.search_timer.setSingleShot(True)
        self.search_timer.setInterval(200)  # 200ms debounce
        self.search_timer.timeout.connect(self.do_search)

        search_row = QHBoxLayout()
        self.search_box = QLineEdit()
        self.search_box.textChanged.connect(self.on_search_text_changed)
        self.search_box.setPlaceholderText("输入 SKU / 中文名 / 英文名 关键字（模糊搜索）")
        self.search_box.returnPressed.connect(self.do_search)
        btn_search = QPushButton("搜索")
        btn_search.clicked.connect(self.do_search)
        btn_select_all = QPushButton("全选")
        btn_select_all.clicked.connect(lambda: self.set_all_checks(True))
        btn_select_none = QPushButton("全不选")
        btn_select_none.clicked.connect(lambda: self.set_all_checks(False))
        search_row.addWidget(self.search_box, 1)
        search_row.addWidget(btn_search)
        search_row.addWidget(btn_select_all)
        search_row.addWidget(btn_select_none)

        self.sku_list = QListWidget()
        self.sku_list.setMinimumHeight(200)

        prod_layout.addLayout(search_row)
        prod_layout.addWidget(QLabel("搜索结果（勾选要导出的SKU）："))
        prod_layout.addWidget(self.sku_list)
        self.grp_prod.setLayout(prod_layout)
        layout.addWidget(self.grp_prod)

        # 导出
        btn_export = QPushButton("生成文件")
        btn_export.clicked.connect(self.do_export)
        layout.addWidget(btn_export)

        self.setLayout(layout)

        # 事件绑定
        self.zone.currentIndexChanged.connect(self.on_zone_changed)
        self.mode.currentIndexChanged.connect(self.on_mode_changed)

        self.on_mode_changed()
        self.on_zone_changed()

    def on_mode_changed(self):
        is_range = (self.mode.currentText() == "库位范围盘点")
        self.grp_range.setVisible(is_range)
        self.grp_prod.setVisible(not is_range)

        # zone：范围模式只允许 C/D；产品模式允许 C区+D区
        current = self.zone.currentText()
        self.zone.blockSignals(True)
        self.zone.clear()
        if is_range:
            self.zone.addItems(["C区", "D区"])
            if current in ("C区", "D区"):
                self.zone.setCurrentText(current)
            else:
                self.zone.setCurrentText("C区")
        else:
            self.zone.addItems(["C区+D区", "C区", "D区"])
            if current in ("C区+D区", "C区", "D区"):
                self.zone.setCurrentText(current)
            else:
                self.zone.setCurrentText("C区+D区")
        self.zone.blockSignals(False)

        self.on_zone_changed()

    def on_zone_changed(self):
        z = self.zone.currentText()
        is_c = (z == "C区")
        is_d = (z == "D区")
        is_range = (self.mode.currentText() == "库位范围盘点")

        # 行范围仅C区有效（范围模式）
        self.r_min.setEnabled(is_c and is_range)
        self.r_max.setEnabled(is_c and is_range)

        if not is_range:
            return

        # ===== 库位范围盘点：切换区域时，默认刷新为最大范围 =====
        self.l_min.setRange(1, 6)
        self.l_max.setRange(1, 6)
        self.l_min.setValue(1)
        self.l_max.setValue(6)

        if is_c:
            self.p_min.setRange(1, 24)
            self.p_max.setRange(1, 24)
            self.p_min.setValue(1)
            self.p_max.setValue(24)

            self.r_min.setRange(1, 48)
            self.r_max.setRange(1, 48)
            self.r_min.setValue(1)
            self.r_max.setValue(48)

        elif is_d:
            self.p_min.setRange(1, 35)
            self.p_max.setRange(1, 35)
            self.p_min.setValue(1)
            self.p_max.setValue(35)

            # 行对D区无效，但为了“看起来一致”，也可以重置成最大（控件已禁用）
            self.r_min.setRange(1, 48)
            self.r_max.setRange(1, 48)
            self.r_min.setValue(1)
            self.r_max.setValue(48)

    def pick_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择WMS导出的库存Excel", "", "Excel Files (*.xlsx *.xls)")
        if not path:
            return
        try:
            df = pd.read_excel(path)
            col = detect_columns(df)
        except Exception as e:
            QMessageBox.critical(self, "读取失败", str(e))
            return

        self.excel_path = path
        self.df = df
        self.col = col
        self.file_label.setText(path)
        self.sku_list.clear()

    def on_search_text_changed(self, _text: str):
        # 没加载Excel就不触发
        if self.df is None or self.col is None:
            return

        # 空输入：清空列表（也可以选择不清空）
        if not self.search_box.text().strip():
            self.sku_list.clear()
            return

        # 重置计时：用户继续输入就不会立刻搜索
        self.search_timer.start()

    def do_search(self):
        if self.df is None or self.col is None:
            QMessageBox.warning(self, "提示", "请先选择Excel文件。")
            return

        q = self.search_box.text().strip()
        if not q:
            self.sku_list.clear()
            return

        df = self.df
        col = self.col

        # 模糊匹配：SKU / CN / EN
        q_low = q.lower()
        sku_s = df[col["sku"]].astype(str)
        cn_s = df[col["cn"]].astype(str)
        en_s = df[col["en"]].astype(str)

        mask = sku_s.str.lower().str.contains(q_low, na=False) | \
               cn_s.str.lower().str.contains(q_low, na=False) | \
               en_s.str.lower().str.contains(q_low, na=False)

        hits = df[mask].copy()
        if hits.empty:
            QMessageBox.information(self, "无结果", "没有匹配到任何SKU。")
            self.sku_list.clear()
            return

        # 取 SKU 唯一项，并展示中/英名（取第一条）
        hits["_sku"] = hits[col["sku"]].astype(str)
        uniq = hits.drop_duplicates(subset=["_sku"])[["_sku", col["cn"], col["en"]]]

        self.sku_list.clear()
        for _, r in uniq.iterrows():
            sku = str(r["_sku"])
            cn = "" if pd.isna(r[col["cn"]]) else str(r[col["cn"]])
            en = "" if pd.isna(r[col["en"]]) else str(r[col["en"]])

            text = f"{sku} | {cn} | {en}"
            item = QListWidgetItem(text)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            # 存SKU
            item.setData(Qt.UserRole, sku)
            self.sku_list.addItem(item)

    def set_all_checks(self, checked: bool):
        for i in range(self.sku_list.count()):
            it = self.sku_list.item(i)
            it.setCheckState(Qt.Checked if checked else Qt.Unchecked)

    def get_selected_skus(self) -> List[str]:
        skus = []
        for i in range(self.sku_list.count()):
            it = self.sku_list.item(i)
            if it.checkState() == Qt.Checked:
                skus.append(str(it.data(Qt.UserRole)))
        return skus
    
    def build_default_pdf_name(self, mode: str, zone_code: str) -> str:
        now = datetime.now().strftime("%Y%m%d_%H%M")

        # 选择一个默认保存目录：优先 Excel 所在目录，否则用户主目录
        base_dir = os.path.dirname(self.excel_path) if self.excel_path else os.path.expanduser("~")

        if mode == "库位范围盘点":
            pmin, pmax = self.p_min.value(), self.p_max.value()
            lmin, lmax = self.l_min.value(), self.l_max.value()

            if zone_code == "C":
                rmin, rmax = self.r_min.value(), self.r_max.value()
                fname = f"C_{pmin}-{pmax}_R{rmin}-{rmax}_L{lmin}-{lmax}_{now}.pdf"
            elif zone_code == "D":
                fname = f"D_{pmin}-{pmax}_L{lmin}-{lmax}_{now}.pdf"
            else:
                fname = f"Inventory_{now}.pdf"
        else:
            selected_count = len(self.get_selected_skus())
            zone_text = "C+D" if zone_code == "ALL" else zone_code
            fname = f"BySKU_{zone_text}_SKU{selected_count}_{now}.pdf"

        return os.path.join(base_dir, fname)

    def do_export(self):
        if self.df is None or self.col is None:
            QMessageBox.warning(self, "提示", "请先选择Excel文件。")
            return

        mode = self.mode.currentText()
        z = self.zone.currentText()

        zone_code = "ALL"
        if z == "C区":
            zone_code = "C"
        elif z == "D区":
            zone_code = "D"

        default_path = self.build_default_pdf_name(mode, zone_code)
        export_kind = self.export_type.currentText()  # "PDF" 或 "Excel"

        # 生成默认文件名后缀：pdf / xlsx
        if export_kind == "Excel" and default_path.lower().endswith(".pdf"):
            default_path = default_path[:-4] + ".xlsx"

        if export_kind == "PDF":
            filter_str = "PDF Files (*.pdf)"
            title_dlg = "保存PDF"
        else:
            filter_str = "Excel Files (*.xlsx)"
            title_dlg = "保存Excel"

        out_path, _ = QFileDialog.getSaveFileName(self, title_dlg, default_path, filter_str)
        if not out_path:
            return

        if export_kind == "PDF" and not out_path.lower().endswith(".pdf"):
            out_path += ".pdf"
        if export_kind == "Excel" and not out_path.lower().endswith(".xlsx"):
            out_path += ".xlsx"

        try:
            if mode == "库位范围盘点":
                pmin, pmax = self.p_min.value(), self.p_max.value()
                lmin, lmax = self.l_min.value(), self.l_max.value()

                if zone_code == "C":
                    rmin, rmax = self.r_min.value(), self.r_max.value()
                    bins = c_aisle_order(pmin, pmax, rmin, rmax, lmin, lmax)
                    title = f"C区盘库表  排{pmin}-{pmax} 行{rmin}-{rmax} 层{lmin}-{lmax}"
                elif zone_code == "D":
                    bins = d_order(pmin, pmax, lmin, lmax)
                    title = f"D区盘库表  排{pmin}-{pmax} 层{lmin}-{lmax}"
                else:
                    raise ValueError("库位范围盘点模式下区域只能选 C区 或 D区。")

                rows = build_rows_for_range(self.df, self.col, bins, zone_code)

            else:
                # 按产品盘点：不补空
                selected = self.get_selected_skus()
                if not selected:
                    QMessageBox.warning(self, "提示", "请先搜索并勾选至少一个SKU。")
                    return

                rows = build_rows_for_products(self.df, self.col, selected, zone_code)
                zone_text = "C区+D区" if zone_code == "ALL" else ("C区" if zone_code == "C" else "D区")
                title = f"按产品盘库表  SKU数={len(selected)}  区域={zone_text}"

                if not rows:
                    QMessageBox.information(self, "无数据", "选中的SKU在所选区域内没有可导出的库存记录。")
                    return

            if export_kind == "PDF":
                export_pdf(rows, out_path, title)
            else:
                export_excel(rows, out_path, title)


        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))
            return

        QMessageBox.information(self, "完成", f"已生成：\n{out_path}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = Main()
    w.setMinimumSize(900, 750)
    w.adjustSize()
    w.show()
    sys.exit(app.exec())
